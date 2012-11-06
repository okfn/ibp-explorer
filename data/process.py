from openpyxl import load_workbook
import json
import re
from string import uppercase

# ######
# Public
# ######

def build_dict(iso_file, q_xls, g_xls, a_xls):
    # Output to populate
    out = {}
    # Workbooks to read
    print 'Reading %s...' % q_xls
    q_workbook = load_workbook(filename=q_xls)
    print 'Reading %s...' % g_xls
    g_workbook = load_workbook(filename=g_xls)
    print 'Reading %s...' % a_xls
    a_workbook = load_workbook(filename=a_xls)
    print 'Reading %s...' % iso_file
    iso_data = json.load(open(iso_file))
    # Generate output
    out['country']  = _read_answers(a_workbook, iso_data)
    out['question'] = _read_questions(q_workbook)
    out['grouping'] = _read_groupings(g_workbook)
    return out


# #########
# Internals
# #########

# Handy lookup of excel column names Don't you love comprehensions?
_COLUMN_NAME = [ (a+b).strip() for a in ' '+uppercase for b in uppercase ]
def _lookup(sheet,x,y):
    """Look up an Excel cell in a sheet. Top left corner is (1,1) not (0,0).
       This avoids remapping rows, as Excel obviously indexes from 1."""
    cell_name = _COLUMN_NAME[x-1] + str(y)
    return sheet.cell(cell_name).value

def _read_answers(a_workbook, iso_data):
    sheet = a_workbook.get_sheet_by_name(name='Sheet1')
    answers = {}
    height = sheet.get_highest_row()
    width = sheet.get_highest_column()
    # Not unlike CSV Dictreader. Also, I like comprehensions.
    headers = { x : _lookup(sheet,x,1)
                for x in range(1,width+1) }
    rows = []
    for y in range(2,height+1):
        row = { headers[x]: _lookup(sheet,x,y)
                for x in range(1,width+1) } 
        rows.append(row)
    # Verify the data
    question_label = re.compile('^q[0-9]+l?$')
    answers = {}
    for row in rows:
        name = row['country']
        year = row['year']
        # Don't trust spreadsheets
        assert type(name) is unicode, '[%s/%s] Invalid country name %s' % (name,year,unicode(name))
        name = name.strip()
        assert name in iso_data, '[%s/%s] I have no ISO-3116 mapping for country name "%s". Please add one to the ISO mappings file.' % (name,year,name)
        assert type(year) is int, '[%s/%s] Invalid year %s (%s)' % (name,year,unicode(year),type(year))
        assert year in [2006,2008,2010,2012], '[%s/%s] Unexpected value of "year": %s' % (name,year,year)
        # Validate the row content
        validated = {}
        for key,value in row.items():
            if question_label.match(key) is not None:
                error_string = '[%s/%s] Invalid value for %s: %s'% (name,year,key,value)
                # Verify the value
                if key[-1]=='l':
                    assert value in [None,'a','b','c','d','e'], error_string
                else:
                    if value is None:
                        # Blank becomes -1
                        value = -1
                    assert type(value) is int, error_string
                    assert value in [100,67,33,0,-1], error_string
                key = key[1:]
            validated[key] = value
        # Store an object grouped by country
        alpha2 = iso_data[name]
        answers[alpha2] = answers.get(alpha2,{ 'name':name, 'alpha2':alpha2 })
        answers[alpha2]['db_%d'%year] = validated
    return sorted(answers.values(),key=lambda x: x['name'])

def _read_questions(q_workbook):
    # Question dict
    sheet = q_workbook.get_sheet_by_name(name='Sheet2')
    questions = {}
    height = sheet.get_highest_row()
    for n in range(2,height+1):
        questions[n-1] = { 
          'number': n,
          'text': _lookup(sheet,2,n),
          'a': _lookup(sheet,3,n),
          'b': _lookup(sheet,4,n),
          'c': _lookup(sheet,5,n),
          'd': _lookup(sheet,6,n),
          'e': _lookup(sheet,7,n),
          }
    return questions

def _read_groupings(g_workbook):
    # Question groupings
    sheet = g_workbook.get_sheet_by_name(name='QuestionsGroups')
    out = []
    # Scroll down column B
    height = sheet.get_highest_row()
    y = 2
    while y <= height:
        title = _lookup(sheet,2,y)
        if title:
            group = {
                'by': title,
                'entries': [],
            }
            while y<height and _lookup(sheet,2,y+1) is not None:
                y += 1
                group['entries'].append({
                  'title': _lookup(sheet,2,y),
                  'qs': _parse_int_list( _lookup(sheet,3,y) ),
                })
            out.append( group )
        y += 1
    return out

def _parse_int_list(int_list):
    if int_list is '' or int_list is u'':
        return []
    if type(int_list) is str:
        int_list = unicode(int_list)
    if (type(int_list) is float) or (type(int_list) is int):
        int_list = unicode(int(int_list))
    
    out = []
    for s in int_list.replace(' ','').split(','):
        split = s.split('-')
        assert len(split)<3, 'Invalid range: %s' % s
        if len(split)==1:
            out.append(int(split[0]))
        else:
            for i in range(int(split[0]), int(split[1])+1):
                out.append(i)
    return out


# TODO
def get_regions(groupings_xls):
    wb = xlrd.open_workbook(groupings_xls)
    sheet = wb.sheet_by_name('CountriesRegions')
    header_row = 2
    out = {}
    for col_number in range(sheet.ncols):
        col = sheet.col_slice(col_number,header_row)
        l = [ col[i].value for i in range(1, len(col)) ]
        # Strip empty strings and store the list
        out[ col[0].value ] = filter(bool,l)
    return out


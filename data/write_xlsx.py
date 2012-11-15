import openpyxl
import unicodecsv

filename_xlsx = 'tmp.xlsx'
filename_csv = 'tmp_%s.csv'

def write(dataset, iso_data):
    # Import all data
    print 'Importing all data...'
    Q_HEADERS, Q_DATA = _get_questions(dataset)
    G_HEADERS, G_DATA = _get_groupings(dataset)
    CN_HEADERS, CN_DATA = _get_countrynames(dataset,iso_data)
    R_HEADERS, R_DATA = _get_regions(dataset)
    S_HEADERS, S_DATA = _get_scores(dataset)
    # Dump XLSX file
    print 'Writing XLSX file...'
    from openpyxl import Workbook
    wb = Workbook()
    first_sheet = wb.get_active_sheet()
    write_sheet(wb,'Questions', Q_HEADERS, Q_DATA)
    write_sheet(wb,'Scores', S_HEADERS, S_DATA, verbose=True)
    write_sheet(wb,'Groupings', G_HEADERS, G_DATA)
    write_sheet(wb,'Regions', R_HEADERS, R_DATA)
    write_sheet(wb,'CountryCodes', CN_HEADERS, CN_DATA)
    wb.remove_sheet(first_sheet)
    wb.save(filename_xlsx)
    # Dump CSV files
    print 'Writing CSV file...'
    write_csv(filename_csv % 'questions', Q_HEADERS, Q_DATA)
    write_csv(filename_csv % 'groupings', G_HEADERS, G_DATA)
    write_csv(filename_csv % 'regions', R_HEADERS, R_DATA)
    write_csv(filename_csv % 'countrycodes', CN_HEADERS, CN_DATA)
    write_csv(filename_csv % 'scores', S_HEADERS, S_DATA)

def write_sheet(wb,sheet_name,headers,data,verbose=False):
    # Dump CSV-style headers and data into an XLSX sheet
    sheet = wb.create_sheet()
    sheet.title = sheet_name
    y = 0
    for x in range(len(headers)):
        sheet.cell(row=y,column=x).value = headers[x]
    total = len(data)
    for row in data:
        y += 1
        if verbose:
            print y,'/',total
        for x in range(len(row)):
            sheet.cell(row=y,column=x).value = row[x]

def write_csv(filename, headers, data):
    with open(filename,'w') as f:
        w = unicodecsv.writer(f)
        w.writerow(headers)
        for x in data:
            w.writerow(x)


###################
### Data Extractors
###################

def _get_regions(dataset):
    # Create Regions sheet
    HEADERS = ['REGION_NAME','COUNTRY']
    DATA = []
    for region in dataset['regions']:
        for country in region['contains']:
            DATA.append( [region['name'],country] )
    return HEADERS, DATA

def _get_countrynames(dataset,iso_data):
    countrypairs = sorted( { y:x for x,y in iso_data.items() }.items(), key=lambda x:x[0] )
    HEADERS = ['ISO3116','COUNTRY_NAME']
    DATA = [ list(x) for x in countrypairs ]
    return HEADERS, DATA

def _get_groupings(dataset):
    HEADERS = ['CATEGORY','GROUP','QUESTION']
    DATA = []
    for category in dataset['groupings']:
        for entry in category['entries']:
            for question in entry['qs']:
                DATA.append( [category['by'],entry['title'],question] )
    return HEADERS, DATA

def _get_questions(dataset):
    HEADERS = ['NUMBER','TEXT','A','B','C','D','E']
    DATA = []
    for x in sorted( dataset['question'].values(), key=lambda x:x['number']):
        DATA.append( [ x['number'],x['text'],x['a'],x['b'],x['c'],x['d'],x['e'] ]  )
    return HEADERS, DATA

def _get_scores(dataset):
    q = range(1,125)
    HEADERS = ['COUNTRY','YEAR']
    for x in q: HEADERS.append(str(x))
    for x in q: HEADERS.append(str(x)+'l')
    DATA = []
    for country in dataset['country']:
        for year in [2006,2008,2010,2012]:
            db = country.get('db_%d'%year)
            if not db: continue
            row = [country['alpha2'],year]
            for x in q:
                row.append( db[str(x)] )
            for x in q:
                score = db[str(x)] 
                letter = {
                        0:'d',
                        -1: 'e',
                        33:'c',
                        67: 'b',
                        100:'a'
                        }[score]
                row.append(letter)
            DATA.append(row)
    return HEADERS, DATA

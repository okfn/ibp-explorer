from openpyxl import load_workbook
import json
import re
from string import uppercase

# ######
# Public
# ######

def read(iso_data, datafiles, old):
    """Read all the input Excel files and produce an enormous data dict."""
    # Output to populate
    dataset = {}
    # Workbooks to read
    print 'Reading %s...' % datafiles['q_xlsx']
    q_workbook = load_workbook(filename=datafiles['q_xlsx'])
    print 'Reading %s...' % datafiles['g_xlsx']
    g_workbook = load_workbook(filename=datafiles['g_xlsx'])
    print 'Reading %s...' % datafiles['a_xlsx']
    a_workbook = load_workbook(filename=datafiles['a_xlsx'])
    print 'Reading %s...' % datafiles['av_xlsx']
    av_workbook = load_workbook(filename=datafiles['av_xlsx'])
    # Generate dataset
    if old:
        dataset['country_old']  = _read_answers(a_workbook, iso_data, datafiles['a_xlsx_sheet'], datafiles['years'])
        dataset['question_old'] = _read_questions(q_workbook, datafiles['q_xlsx_sheet'])
        dataset['groupings_old'] = _read_groupings(g_workbook, datafiles['g_xlsx_qsheet'])
        dataset['regions_old'] = _read_regions(g_workbook, iso_data, datafiles['g_xlsx_csheet'])
        dataset['availability_old'] = _read_availability(av_workbook, iso_data, datafiles['av_xlsx_sheets'])
    else:
        print 'Reading %s...' % datafiles['pp_xlsx']
        pp_workbook = load_workbook(filename=datafiles['pp_xlsx'])
        dataset['country']  = _read_answers(a_workbook, iso_data, datafiles['a_xlsx_sheet'], datafiles['years'])
        dataset['question'] = _read_questions(q_workbook, datafiles['q_xlsx_sheet'])
        dataset['groupings'] = _read_groupings(g_workbook, datafiles['g_xlsx_qsheet'])
        dataset['regions'] = _read_regions(g_workbook, iso_data, datafiles['g_xlsx_csheet'])
        dataset['availability'] = _read_availability(av_workbook, iso_data, datafiles['av_xlsx_sheets'])
        dataset['public_participation'] = _read_participation(pp_workbook, iso_data, datafiles['pp_xlsx_sheet'])
    return dataset


# #########
# Internals
# #########

# Handy lookup of excel column names Don't you love comprehensions?
_COLUMN_NAME = [ (a+b).strip() for a in ' '+uppercase for b in uppercase ]
def _lookup(sheet,x,y):
    """Look up an Excel cell in a sheet. Top left corner is (1,1) not (0,0).
       This avoids remapping rows, as Excel obviously indexes from 1."""
    cell_name = _COLUMN_NAME[x-1] + str(y)
    return sheet.cell(cell_name).value

def _read_answers(a_workbook, iso_data, sheet_name, years):
    sheet = a_workbook.get_sheet_by_name(name=sheet_name)
    answers = {}
    height = sheet.get_highest_row()
    width = sheet.get_highest_column()
    # Not unlike CSV Dictreader. Also, I like comprehensions.
    headers = { x : _lookup(sheet,x,1)
                for x in range(1,width+1) }
    rows = []
    for y in range(2,height+1):
        row = { headers[x]: _lookup(sheet,x,y)
                for x in range(1,width+1) } 
        rows.append(row)
    # Verify the data
    question_label = re.compile('^((q[0-9]+l?)|(t3[a-z]{2,3}l?))$')
    answers = {}
    for row in rows:
        name = row['country']
        year = row['year']
        # Don't trust spreadsheets
        assert type(name) is unicode, '[%s/%s] Invalid country name %s' % (name,year,unicode(name))
        name = name.strip()
        assert name in iso_data, '[%s/%s] I have no ISO-3116 mapping for country name "%s". Please add one to the ISO mappings file.' % (name,year,name)
        assert type(year) is int, '[%s/%s] Invalid year %s (%s)' % (name,year,unicode(year),type(year))
        assert year in years, '[%s/%s] Unexpected value of "year": %s' % (name,year,year)
        # Validate the row content
        validated = {}
        for key,value in row.items():
            if key is None: continue
            if question_label.match(key) is not None:
                error_string = '[%s/%s] Invalid value for %s: %s'% (name,year,key,value)
                # Verify the value
                if key[-1]=='l':
                    assert value in [None,'a','b','c','d','e'], error_string
                else:
                    if value is None:
                        # Blank becomes -1
                        value = -1
                    assert type(value) is int, error_string
                    assert value in [100,67,33,0,-1], error_string
                if re.search('^q[0-9]+l?$', key):
                    key = key[1:]
                elif re.search('^t3[a-z]{2,3}l?$', key):
                    key = key
            validated[key] = value
        # Store an object grouped by country
        alpha2 = iso_data[name]
        answers[alpha2] = answers.get(alpha2,{ 'name':name, 'alpha2':alpha2 })
        answers[alpha2]['db_%d'%year] = validated
    return sorted(answers.values(),key=lambda x: x['name'])

def _read_questions(q_workbook, sheet_name):
    # Question dict
    sheet = q_workbook.get_sheet_by_name(name=sheet_name)
    questions = {}
    height = sheet.get_highest_row()
    for n in range(2,height+1):
        questions[n-1] = { 
          'number': _lookup(sheet,1,n),
          'text': _lookup(sheet,2,n),
          'a': _lookup(sheet,3,n),
          'b': _lookup(sheet,4,n),
          'c': _lookup(sheet,5,n),
          'd': _lookup(sheet,6,n),
          'e': _lookup(sheet,7,n),
          }
    # Fix '(Please comment)' text appearing in most answers
    regex = r'\s?\(Please comment\)'
    for q in questions.values():
        for key in ['a','b','c','d','e']:
            if q[key] is None: continue
            q[key] = re.sub( regex,'',q[key], flags=re.IGNORECASE )
    return questions

def _read_groupings(g_workbook, sheet_name):
    # Question groupings
    sheet = g_workbook.get_sheet_by_name(name=sheet_name)
    out = []
    # Scroll down column B
    height = sheet.get_highest_row()
    y = 2
    while y <= height:
        title = _lookup(sheet,2,y)
        if title:
            group = {
                'by': title,
                'entries': [],
            }
            while y<height and _lookup(sheet,2,y+1) is not None:
                y += 1
                group['entries'].append({
                  'title': _lookup(sheet,2,y),
                  'qs': _parse_int_list( _lookup(sheet,3,y) ),
                })
            out.append( group )
        y += 1
    return out

def _read_regions(g_workbook, iso_data, sheet_name):
    # Question groupings
    sheet = g_workbook.get_sheet_by_name(name=sheet_name)
    out = []
    # Scroll along row 3
    x = 1
    while True:
        title = _lookup(sheet,x,3)
        if not title: break
        region = {
                'name': title,
                'contains': []
                }
        # Scroll down the remainder of this column
        y = 4
        while True:
            country_name = _lookup(sheet,x,y)
            if not country_name: break
            assert country_name in iso_data, '[Groupings/%s] I have no ISO-3116 mapping for country name "%s". Please add one to the ISO mappings file.' % (sheet_name,country_name)
            region['contains'].append(iso_data[country_name])
            y+=1
        out.append(region)
        x += 1
    return out

def _read_availability(av_workbook, iso_data, sheets):
    column_headers = [('country', 'country'),
            ('Pre-Budget Statement', 'prebudgetstatement'),
            ('Executive\'s Budget Proposal', 'executivesbudgetproposal'),
            ('Citizens Budget', 'citizensbudget'),
            ('Enacted Budget', 'enactedbudget'),
            ('In-Year Reports', 'inyearreports'),
            ('Mid-Year Review', 'midyearreview'),
            ('Year-End Report', 'yearendreport'),
            ('Audit Report', 'auditreport')]
    valid = ['PW', 'IU', 'NP', 'HC', 'SC']
    out = {}
    for sheet_name in sheets:
        sheet = av_workbook.get_sheet_by_name(name=sheet_name)
        # Check the sheet is structured the way we expect
        for x in range(1,len(column_headers)+1):
            column_header = _lookup(sheet, x,1) 
            assert column_header == column_headers[x-1][0], 'Bad column header on sheet %s: Got "%s" but expected "%s".' % (sheet_name,column_header,column_headers[x-1])
        # Scraped the data into a structured JSON object
        height = sheet.get_highest_row()
        for y in range(2,height+1):
            name = _lookup(sheet,1,y)
            assert name in iso_data, '[%s/%s] I have no ISO-3116 mapping for country name "%s". Please add one to the ISO mappings file.' % (name,year,name)
            alpha2 = iso_data[name]
            data = {}
            data['alpha2'] = alpha2
            data['name'] = name
            for x in range(2,len(column_headers)+1): # <-- Mind these off-by-ones! x is column number not list index.
                value = _lookup(sheet,x,y)
                assert value in valid, 'Invalid value in %s (%s, %s): %s' % (sheet_name,name,column_headers[x-1],value)
                data[column_headers[x-1][1]] = value
            # Write the data into the structured JSON object
            out[alpha2] = out.get(alpha2,{})
            out[alpha2]['db_'+sheet_name] = data
    return sorted(out.values(), key=lambda x:x.values()[0]['name'])

def _read_participation(pp_workbook, iso_data, sheet_name):
    sheet = pp_workbook.get_sheet_by_name(name=sheet_name)
    participation = []
    vtypes = {'Score': 'score', 'Letter': 'letter', 'Comments and Citations in support of the response': 'comments'}
    height = sheet.get_highest_row()
    width = sheet.get_highest_column()

    headers = { x : (_lookup(sheet,x,1), _lookup(sheet,x,2))
                for x in range(1,width+1) }
    for y in range(3,height+1):
        name = _lookup(sheet,1,y).strip()
        assert name in iso_data, '[%s] I have no ISO-3116 mapping for country name "%s". Please add one to the ISO mappings file.' % (name,name)
        alpha2 = iso_data[name]
        data = {}
        data['alpha2'] = alpha2
        data['name'] = name
        for x in range(2,width+1):
            value = _lookup(sheet,x,y)
            if not value:
                value = ''
            value = value.strip()
            question = headers[x][0].strip()
            vtype = headers[x][1].strip()
            vkey = vtypes[vtype]
            if vkey == 'letter':
                value = value.replace('.','')
            if question in data:
                data[question][vkey] = value
            else:
                data[question] = {}
                data[question][vkey] = value
        participation.append(data)
    return participation

def _parse_int_list(int_list):
    if int_list is '' or int_list is u'':
        return []
    if type(int_list) is str:
        int_list = unicode(int_list)
    if (type(int_list) is float) or (type(int_list) is int):
        int_list = unicode(int(int_list))
    
    out = []
    for s in int_list.replace(' ','').split(','):
        split = s.split('-')
        assert len(split)<3, 'Invalid range: %s' % s
        if len(split)==1:
            out.append(split[0])
        else:
            for i in range(int(split[0]), int(split[1])+1):
                out.append(str(i))
    return out


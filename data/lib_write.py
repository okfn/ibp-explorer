import unicodecsv
import os
import json
import zipfile


def write_js(dataset, jsonfilename):
    with open(jsonfilename, 'w') as jsonfile:
        output_js = 'window._EXPLORER_DATASET = %s;' % json.dumps(dataset)
        print >>jsonfile, output_js
    print('wrote %s' % jsonfilename)


def write_downloads(dataset, iso_data, downloadfoldername, files, years):
    # Populate the downloads folder
    # -----------------------------
    if years[0] >= 2015:
        keys = {}
        keys['question'] = 'question_%s' % years[0]
        keys['country'] = 'country_%s' % years[0]
        keys['groupings'] = 'groupings_%s' % years[0]
        keys['regions'] = 'regions_%s' % years[0]
        keys['availability'] = 'availability_%s' % years[0]
        keys['downloads'] = 'downloads_%s' % years[0]
    else:
        keys = {}
        keys['question'] = 'question_old'
        keys['country'] = 'country_old'
        keys['groupings'] = 'groupings_old'
        keys['regions'] = 'regions_old'
        keys['availability'] = 'availability_old'
        keys['downloads'] = 'downloads_old'
    # Import all data
    print('Importing all data...')
    Q_HEADERS, Q_DATA = _questions_as_csv(dataset, keys['question'])
    G_HEADERS, G_DATA = _groupings_as_csv(dataset, keys['groupings'])
    CN_HEADERS, CN_DATA = _countrynames_as_csv(dataset, iso_data)
    R_HEADERS, R_DATA = _regions_as_csv(dataset, keys['regions'])
    S_HEADERS, S_DATA = _scores_as_csv(dataset, years, keys['country'])
    SM_HEADERS, SM_DATA = _summary_as_csv(dataset, years,
                                          keys['country'], keys['question'])
    if years[0] == 2015:
        P_HEADERS, P_DATA = participation_as_csv(dataset)
    # Dump XLSX file
    print('Writing XLSX file...')
    from openpyxl import Workbook
    wb = Workbook()
    first_sheet = wb.active
    _write_sheet(wb, 'Summary', SM_HEADERS, SM_DATA)
    _write_sheet(wb, 'Questions', Q_HEADERS, Q_DATA)
    _write_sheet(wb, 'Scores', S_HEADERS, S_DATA)
    _write_sheet(wb, 'Groupings', G_HEADERS, G_DATA)
    if years[0] == 2015:
        _write_sheet(wb, 'PublicParticipation', P_HEADERS, P_DATA)
    _write_sheet(wb, 'Regions', R_HEADERS, R_DATA)
    _write_sheet(wb, 'CountryCodes', CN_HEADERS, CN_DATA)
    wb.remove_sheet(first_sheet)
    wb.save(os.path.join(downloadfoldername, files['xlsx']))
    # Dump JSON file
    with open(os.path.join(downloadfoldername, files['json']),
              'w') as jsonfile:
        json.dump(dataset, jsonfile)
    # Dump CSV files
    print('Writing CSV file...')
    csv_q = files['csv'] % 'questions'
    csv_g = files['csv'] % 'groupings'
    csv_r = files['csv'] % 'regions'
    csv_c = files['csv'] % 'countrycodes'
    csv_s = files['csv'] % 'scores'
    csv_sm = files['csv'] % 'summary'
    if years[0] == 2015:
        csv_p = files['csv'] % 'public_participation'
        _write_csv(csv_p, P_HEADERS, P_DATA)
    _write_csv(csv_q, Q_HEADERS, Q_DATA)
    _write_csv(csv_g, G_HEADERS, G_DATA)
    _write_csv(csv_r, R_HEADERS, R_DATA)
    _write_csv(csv_c, CN_HEADERS, CN_DATA)
    _write_csv(csv_s, S_HEADERS, S_DATA)
    _write_csv(csv_sm, SM_HEADERS, SM_DATA)
    # Create zip file
    print('Writing ZIP file...')
    with zipfile.ZipFile(os.path.join(downloadfoldername, files['csv_zip']),
                         'w') as z:
        z.write(csv_q)
        z.write(csv_g)
        z.write(csv_r)
        z.write(csv_c)
        z.write(csv_s)
        z.write(csv_sm)
        if years[0] == 2015:
            z.write(csv_p)
    os.unlink(csv_q)
    os.unlink(csv_g)
    os.unlink(csv_r)
    os.unlink(csv_c)
    os.unlink(csv_s)
    os.unlink(csv_sm)
    if years[0] == 2015:
        os.unlink(csv_p)
    # Create list of downloads
    downloads = [
        {'filename': files['xlsx'], 'format': 'Excel'},
        {'filename': files['csv_zip'], 'format': 'CSV'},
        {'filename': files['json'], 'format': 'JSON'}
    ]
    # Mapping of filenames
    for x in downloads:
        filename = os.path.join(downloadfoldername, x['filename'])
        raw_size = os.path.getsize(filename)
        x['size'] = _format_kilobytes(raw_size)
    dataset[keys['downloads']] = downloads
    return dataset


##################
# Helper Utilities
##################
def _format_kilobytes(x):
    x = str(x/1024)
    if len(x) > 3:
        x = x[:-3]+','+x[-3:]
    return x+'kb'


def _write_sheet(wb, sheet_name, headers, data):
    # Dump CSV-style headers and data into an XLSX sheet
    sheet = wb.create_sheet()
    sheet.title = sheet_name
    y = 1
    for x in range(len(headers)):
        sheet.cell(row=y, column=x+1).value = headers[x]
    total = len(data)
    for row in data:
        y += 1
        for x in range(len(row)):
            sheet.cell(row=y, column=x+1).value = row[x]


def _write_csv(filename, headers, data):
    with open(filename, 'w') as f:
        w = unicodecsv.writer(f)
        w.writerow(headers)
        for x in data:
            w.writerow(x)


###################
# Data Extractors
###################

def _regions_as_csv(dataset, regions):
    # Create Regions sheet
    HEADERS = ['REGION_NAME', 'COUNTRY_CODE']
    DATA = []
    for region in dataset[regions]:
        for country in region['contains']:
            DATA.append([region['name'], country])
    return HEADERS, DATA


def _countrynames_as_csv(dataset, iso_data):
    countrypairs = sorted({y: x for x, y in iso_data.items()}.items(),
                          key=lambda x: x[0])
    HEADERS = ['COUNTRY_CODE', 'COUNTRY_NAME']
    DATA = [list(x) for x in countrypairs]
    return HEADERS, DATA


def _groupings_as_csv(dataset, groupings):
    HEADERS = ['CATEGORY', 'GROUP', 'QUESTION']
    DATA = []
    t3q = {'134': 't3pbs', '135': 't3ebp', '136': 't3eb', '137': 't3iyr',
           '138': 't3myr', '139': 't3yer', '140': 't3ar'}
    for category in dataset[groupings]:
        for entry in category['entries']:
            for question in entry['qs']:
                if question in t3q:
                    q = t3q[question]
                else:
                    q = question
                DATA.append([category['by'], entry['title'], q])
    return HEADERS, DATA


def _questions_as_csv(dataset, questions):
    HEADERS = ['NUMBER', 'TEXT', 'A', 'B', 'C', 'D', 'E']
    DATA = []
    for x in sorted(dataset[questions].values(), key=lambda x: x['number']):
        DATA.append([x['number'], x['text'], x['a'],
                     x['b'], x['c'], x['d'], x['e']])
    return HEADERS, DATA


def _scores_as_csv(dataset, years, countries):
    if years[0] == 2015:
        q = range(1, 141)
        HEADERS = ['COUNTRY_CODE', 'YEAR']
        t3q = {134: 't3pbs', 135: 't3ebp', 136: 't3eb', 137: 't3iyr',
               138: 't3myr', 139: 't3yer', 140: 't3ar'}
        for x in q:
            if x >= 134:
                HEADERS.append(t3q[x])
            else:
                HEADERS.append('Q '+str(x))
        for x in q:
            if x >= 134:
                HEADERS.append(t3q[x] + ' (LETTER)')
            else:
                HEADERS.append('Q '+str(x)+' (LETTER)')
    elif years[0] == 2017:
        q = range(1, 150)
        HEADERS = ['COUNTRY_CODE', 'YEAR']
        t3q = {143: 'PBS-2', 144: 'EBP-2', 145: 'EB-2', 146: 'IYR-2',
               147: 'MYR-2', 148: 'YER-2', 149: 'AR-2'}
        for x in q:
            if x >= 143:
                HEADERS.append(t3q[x])
            else:
                HEADERS.append('Q '+str(x))
        for x in q:
            if x >= 143:
                HEADERS.append(t3q[x] + ' (LETTER)')
            else:
                HEADERS.append('Q '+str(x)+' (LETTER)')
    elif years[0] == 2019:
        q = range(1, 150)
        HEADERS = ['COUNTRY_CODE', 'YEAR']
        t3q = {143: 'PBS-2', 144: 'EBP-2', 145: 'EB-2', 146: 'IYRs-2',
               147: 'MYR-2', 148: 'YER-2', 149: 'AR-2'}
        for x in q:
            if x >= 143:
                HEADERS.append(t3q[x])
            else:
                HEADERS.append('Q '+str(x))
        for x in q:
            if x >= 143:
                HEADERS.append(t3q[x] + ' (LETTER)')
            else:
                HEADERS.append('Q '+str(x)+' (LETTER)')
    else:
        HEADERS = ['COUNTRY_CODE', 'YEAR']
        q = sorted([int(d) for d in dataset[countries][0].get('db_%d' %
                   years[-1]).keys() if unicode.isdigit(d)])  # noqa
        for x in q:
            HEADERS.append('Q '+str(x))
        for x in q:
            HEADERS.append('Q '+str(x)+' (LETTER)')
    DATA = []
    for country in dataset[countries]:
        for year in years:
            db = country.get('db_%d' % year)
            if not db:
                continue
            row = [country['alpha2'], year]
            for x in q:
                if (year == 2015 and x >= 134) or (year in [2017, 2019] and x >= 143):
                    qkey = t3q[x]
                    row.append(db[qkey])
                else:
                    row.append(db[str(x)])
            for x in q:
                if (year == 2015 and x >= 134) or (year in [2017, 2019] and x >= 143):
                    qkey = t3q[x]
                    score = db[qkey]
                else:
                    score = db[str(x)]
                letter = {
                    0: 'd',
                    -1: 'e',
                    33: 'c',
                    67: 'b',
                    100: 'a'
                 }[score]
                row.append(letter)
            # Replace -1 with None for values in row
            row = [None if i == -1 else i for i in row]
            DATA.append(row)
    return HEADERS, DATA


def _summary_as_csv(dataset, years, countries, questions):
    q = range(1, 125)
    HEADERS = ['COUNTRY', 'YEAR', 'OPEN_BUDGET_INDEX', 'RANK']
    DATA = []
    for year in years:
        temp = []
        for country in dataset[countries]:
            db = country.get('db_%d' % year)
            if not db:
                continue
            score = db.get('roundobi')
            row = [country['alpha2'], year, score, -1]
            temp.append(row)
        # Sort this year's array by score
        temp = sorted(temp, key=lambda x: x[2], reverse=True)
        # Add rankings
        rank = 0
        latest = -1
        for i in range(len(temp)):
            if temp[i][2] != latest:
                latest = temp[i][2]
                rank = i+1
            temp[i][3] = rank
        # Fold this year's results into the overall results
        for x in temp:
            DATA.append(x)
    DATA = sorted(DATA, key=lambda x: x[0])
    return HEADERS, DATA


def participation_as_csv(dataset):
    q = range(119, 134)
    HEADERS = ['COUNTRY', 'YEAR', 'QUESTION', 'SCORE', 'LETTER', 'COMMENTS']
    DATA = []
    for country in dataset['public_participation']:
        for x in q:
            question = country[str(x)]
            row = [country['alpha2'], 2015, str(x), float(question['score']),
                   question['letter'], question['comments']]
            DATA.append(row)
    DATA = sorted(DATA, key=lambda x: x[0])
    return HEADERS, DATA
